from __future__ import print_function, division
import torch
import torch.nn as nn
import torch.optim as optim
from torch.optim import lr_scheduler
import numpy as np
import torchvision
from torchvision import transforms as T
from torchvision import datasets, models
import matplotlib.pyplot as plt
import time
import math
import os
import copy
import random
import torch.utils.data as data
import pickle
from sklearn.metrics import roc_auc_score, f1_score, accuracy_score, recall_score, precision_score, roc_curve, confusion_matrix
from _collections import OrderedDict
import openpyxl
from sklearn.model_selection import KFold, StratifiedKFold
import scipy.stats as stats
from sklearn.utils import resample
from model import Combined_network
import monai
from monai.utils import set_determinism
import torch.nn.functional as F

def log_create():
    info = openpyxl.Workbook()
    all_sheet = info.create_sheet('1', 0)
    all_sheet.cell(1, 1, 'learning_rate')
    all_sheet.cell(1, 2, 'batch_size')
    all_sheet.cell(1, 3, 'epoch_id')
    all_sheet.cell(1, 4, 'loss_train')
    all_sheet.cell(1, 5, 'AUC_IVC')
    all_sheet.cell(1, 6, 'AUC_EVC1')
    all_sheet.cell(1, 7, 'AUC_EVC2')
    all_sheet.cell(1, 8, 'ACC_IVC')
    all_sheet.cell(1, 9, 'ACC_EVC1')
    all_sheet.cell(1, 10, 'ACC_EVC2')
    all_sheet.cell(1, 11, 'F1_IVC')
    all_sheet.cell(1, 12, 'F1_EVC1')
    all_sheet.cell(1, 13, 'F1_EVC2')
    all_sheet.cell(1, 14, 'AUC_IVC_95Cl_r')
    all_sheet.cell(1, 15, 'AUC_IVC_95Cl_l')
    all_sheet.cell(1, 16, 'AUC_EVC1_95Cl_r')
    all_sheet.cell(1, 17, 'AUC_EVC1_95Cl_l')
    all_sheet.cell(1, 18, 'AUC_EVC2_95Cl_r')
    all_sheet.cell(1, 19, 'AUC_EVC2_95Cl_l')
    all_sheet.cell(1, 20, 'ACC_IVC_95Cl_r')
    all_sheet.cell(1, 21, 'ACC_IVC_95Cl_l')
    all_sheet.cell(1, 22, 'ACC_EVC1_95Cl_r')
    all_sheet.cell(1, 23, 'ACC_EVC1_95Cl_l')
    all_sheet.cell(1, 24, 'ACC_EVC2_95Cl_r')
    all_sheet.cell(1, 25, 'ACC_EVC2_95Cl_l')
    all_sheet.cell(1, 26, 'F1_IVC_95_r')
    all_sheet.cell(1, 27, 'F1_IVC_95_l')
    all_sheet.cell(1, 28, 'F1_EVC1_95_r')
    all_sheet.cell(1, 29, 'F1_EVC1_95_l')
    all_sheet.cell(1, 30, 'F1_EVC2_95_r')
    all_sheet.cell(1, 31, 'F1_EVC2_95_l')
    return info

def global_input(current_id, patches):
    if len(patches) == 1:
        return patches[current_id], patches[current_id]
    
    current_patch = patches[current_id]
    patch_size = current_patch.shape
    remaining_patches = [patch for i, patch in enumerate(patches) if i != current_id]
    random.shuffle(remaining_patches)
    grid_size = int(np.ceil(np.sqrt(len(remaining_patches))))
    
    combined_patch = np.zeros((grid_size * patch_size[0], grid_size * patch_size[1], patch_size[2]), dtype=patches[0].dtype)
    for i, patch in enumerate(remaining_patches):
        row = i // grid_size
        col = i % grid_size
        combined_patch[row * patch_size[0]:(row + 1) * patch_size[0],
                       col * patch_size[1]:(col + 1) * patch_size[1], :] = patch
    
    return current_patch, combined_patch


class LN_Dataset(data.Dataset):
    def __init__(self, p_T2_patch_list, p_patch_LD_list, p_patch_SD_list, p_patch_RD_list, p_patch_adc_list, p_LN_meta_label_list, p_LN_meta_ratio):
        self.list_p_patch = p_T2_patch_list
        self.list_p_patch_LD = p_patch_LD_list
        self.list_p_patch_SD = p_patch_SD_list
        self.list_p_patch_RD = p_patch_RD_list
        self.list_p_patch_adc = p_patch_adc_list
        self.list_p_LN_meta_label = p_LN_meta_label_list
        self.list_p_LN_meta_ratio = p_LN_meta_ratio

    def __getitem__(self, idx):
        p_patch = self.list_p_patch[idx]
        p_patch_LD = self.list_p_patch_LD[idx]
        p_patch_SD = self.list_p_patch_SD[idx]
        p_patch_RD = self.list_p_patch_RD[idx]
        p_patch_adc = self.list_p_patch_adc[idx]
        p_LN_meta_label = self.list_p_LN_meta_label[idx]
        p_LN_meta_ratio = self.list_p_LN_meta_ratio[idx]
        return p_patch, p_patch_LD, p_patch_SD, p_patch_RD, p_patch_adc, p_LN_meta_label, p_LN_meta_ratio

    def __len__(self):
        return len(self.list_p_LN_meta_label)

def my_collate(batch):
    p_patches = []
    p_patch_LDs = []
    p_patch_SDs = []
    p_patch_RDs = []
    p_patch_adcs = []
    p_LN_meta_labels = []
    p_LN_meta_ratios = []
    for p_patch, p_patch_LD, p_patch_SD, p_patch_RD, p_patch_adc, p_LN_meta_label, p_LN_meta_ratio in batch:
        p_patches.append(p_patch)
        p_patch_LDs.append(p_patch_LD)
        p_patch_SDs.append(p_patch_SD)
        p_patch_RDs.append(p_patch_RD)
        p_patch_adcs.append(p_patch_adc)
        p_LN_meta_labels.append(p_LN_meta_label)
        p_LN_meta_ratios.append(p_LN_meta_ratio)
    return p_patches, p_patch_LDs, p_patch_SDs, p_patch_RDs, p_patch_adcs, p_LN_meta_labels, p_LN_meta_ratios

def setup_seed(seed):
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    np.random.seed(seed)
    random.seed(seed)
    torch.backends.cudnn.deterministic = True 

def load_LN_data(center_name, data_folder_path):
    all_patches = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/'
                                  + 'T2_2D_patches_correct_3channel.bin', "rb")) 
    all_patch_LDs = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/'
                                  + 'Patches_long_diameter.bin', "rb"))
    all_patch_SDs = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/'
                                  + 'Patches_short_diameter.bin', "rb"))
    all_patch_RDs = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/'
                                  + 'Patches_ratio_diameter.bin', "rb"))
    all_patch_adc = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/'
                                  + 'Patches_adc_value.bin', "rb"))
    all_LN_meta_labels = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/P_label.bin', "rb"))
    all_LN_meta_ratios = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/P_LN_meta_ratio.bin', "rb"))

    ban_list = pickle.load(open(data_folder_path + center_name + '/patches_2d_bin_std_all/P_ban_list.bin', "rb"))
    all_p_ind = list(range(len(all_patches)))
    all_p_include_ind = []
    for ind in all_p_ind:
        if ind not in ban_list:
            all_p_include_ind.append(ind)

    all_patches = [all_patches[i] for i in all_p_include_ind]
    all_patch_LDs = [all_patch_LDs[i] for i in all_p_include_ind]
    all_patch_SDs = [all_patch_SDs[i] for i in all_p_include_ind]
    all_patch_RDs = [all_patch_RDs[i] for i in all_p_include_ind]
    all_patch_adc = [all_patch_adc[i] for i in all_p_include_ind]
    all_LN_meta_labels = [all_LN_meta_labels[i] for i in all_p_include_ind]
    all_LN_meta_ratios = [all_LN_meta_ratios[i] for i in all_p_include_ind]

    return all_patches, all_patch_LDs, all_patch_SDs, all_patch_RDs, all_patch_adc, all_LN_meta_labels, all_LN_meta_ratios

def load_LN_data_by_set(training_data_center_name, data_folder_path, set_name): 
    all_p_patches = pickle.load(open(data_folder_path + training_data_center_name + '/patches_2d_bin_std_all/'
                                     + 'T2_2D_patches_correct_3channel.bin', "rb"))
    all_p_patch_LDs = pickle.load(
        open(data_folder_path + training_data_center_name + '/patches_2d_bin_std_all/Patches_long_diameter.bin', "rb"))
    all_p_patch_SDs = pickle.load(
        open(data_folder_path + training_data_center_name + '/patches_2d_bin_std_all/Patches_short_diameter.bin', "rb"))
    all_p_patch_RDs = pickle.load(
        open(data_folder_path + training_data_center_name + '/patches_2d_bin_std_all/Patches_ratio_diameter.bin', "rb"))
    all_p_patch_adc = pickle.load(
        open(data_folder_path + training_data_center_name + '/patches_2d_bin_std_all/Patches_adc_value.bin', "rb"))
    all_LN_meta_labels = pickle.load(
        open(data_folder_path + training_data_center_name + '/patches_2d_bin_std_all/P_label.bin', "rb"))
    all_LN_meta_ratios = pickle.load(
        open(data_folder_path + training_data_center_name + '/patches_2d_bin_std_all/P_LN_meta_ratio.bin', "rb"))

    set_ind = pickle.load(
        open(data_folder_path + training_data_center_name + '/patches_2d_bin_std_all/' + set_name + '_ind_random_666.bin',
             "rb"))
    set_ind_selected = []
    ban_list = pickle.load(open(data_folder_path + training_data_center_name + '/patches_2d_bin_std_all/P_ban_list.bin', "rb"))
    for ind in set_ind:
        if ind not in ban_list:
            set_ind_selected.append(ind)

    p_patches = [all_p_patches[i] for i in set_ind_selected]
    p_patch_LDs = [all_p_patch_LDs[i] for i in set_ind_selected]
    p_patch_SDs = [all_p_patch_SDs[i] for i in set_ind_selected]
    p_patch_RDs = [all_p_patch_RDs[i] for i in set_ind_selected]
    p_patch_adc = [all_p_patch_adc[i] for i in set_ind_selected]
    p_LN_meta_label = [all_LN_meta_labels[i] for i in set_ind_selected]
    p_LN_meta_ratio = [all_LN_meta_ratios[i] for i in set_ind_selected]

    return p_patches, p_patch_LDs, p_patch_SDs, p_patch_RDs, p_patch_adc, p_LN_meta_label, p_LN_meta_ratio

def calculate_accuracy_ci(y_true, y_pred, confidence=0.95):
    acc = accuracy_score(y_true, y_pred)
    n = len(y_true)
    se = np.sqrt(acc * (1 - acc) / n)
    z = stats.norm.ppf((1 + confidence) / 2)
    margin_of_error = z * se
    return acc, (acc - margin_of_error, acc + margin_of_error)

def calculate_auc_ci(y_true, y_scores, n_iterations=1000, confidence=0.95):
    auc = roc_auc_score(y_true, y_scores)
    aucs = []
    for _ in range(n_iterations):
        indices = resample(np.arange(len(y_true)), replace=True)
        aucs.append(roc_auc_score(y_true[indices], y_scores[indices]))
    lower = np.percentile(aucs, (1 - confidence) * 100 / 2)
    upper = np.percentile(aucs, 100 - (1 - confidence) * 100 / 2)
    return auc, (lower, upper)

def calculate_f1_ci(y_true, y_pred, confidence=0.95):
    f1 = f1_score(y_true, y_pred)
    n = len(y_true)
    se = np.sqrt(f1 * (1 - f1) / n)
    z = stats.norm.ppf((1 + confidence) / 2)
    margin_of_error = z * se
    return f1, (f1 - margin_of_error, f1 + margin_of_error)

def evaluate(y_true, y_pred, digits=3, cutoff='auto'):
    if cutoff == 'auto':
        fpr, tpr, thresholds = roc_curve(y_true, y_pred)
        youden = tpr-fpr
        cutoff = thresholds[np.argmax(youden)]

    y_pred_t = [1 if i > cutoff else 0 for i in y_pred]

    evaluation = OrderedDict()
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred_t).ravel()
    evaluation['AUC'] = round(roc_auc_score(y_true, y_pred), digits)
    evaluation['ACC'] = round(accuracy_score(y_true, y_pred_t), digits)
    evaluation['SEN'] = round(tp / (tp + fn), digits) 
    evaluation['SPE'] = round(tn / (tn + fp), digits)
    evaluation['cutoff'] = cutoff
    evaluation['F1'] = f1_score(y_true, y_pred_t)

    acc, acc_ci = calculate_accuracy_ci(y_true, y_pred_t)
    auc, auc_ci = calculate_auc_ci(y_true, y_pred)
    f1, f1_ci = calculate_f1_ci(y_true, y_pred_t)
    return evaluation, acc_ci, auc_ci, f1_ci


def train_model(device, dataloaders, model, criterions, optimizer, scheduler,
                batch_size, learning_rate, num_epochs, work_dir, log_path, log_modelling, index_record, ra_loss_weight, ram_start):

    patch_h = 8
    patch_w = 8
    transform = T.Compose([
        T.GaussianBlur(9, sigma=(0.1, 2.0)),
        T.Resize((patch_h * 14, patch_w * 14), antialias=True),  
        T.CenterCrop((patch_h * 14, patch_w * 14)),
        T.Normalize(mean=(0.485, 0.456, 0.406), std=(0.229, 0.224, 0.225)),  
    ])
    since = time.time()
    random_transform = T.Compose([
        T.RandomHorizontalFlip(),
        T.RandomVerticalFlip(),
    ])

    results_save_folder = work_dir + 'lr_{}_bs_{}/'.format(learning_rate, batch_size)
    if not os.path.exists(results_save_folder):
        os.makedirs(results_save_folder)

    shenames = log_modelling.get_sheet_names()
    all_sheet = log_modelling[shenames[0]]
    #
    AUC_train = 0.0
    AUC_val = 0.0
    AUC_EVC1 = 0.0
    AUC_EVC2 = 0.0

    best_val_auc = 0.0

    for epoch in range(num_epochs):
        epoch_id = epoch + 1
        print('Epoch {}/{}'.format(epoch_id, num_epochs))
        print('-' * 10)
        print('running learning rate: ', optimizer.param_groups[0]['lr'])
        learning_rate_current = optimizer.param_groups[0]['lr']
        epoch_start_time = time.time()
        all_sheet.cell(index_record, 1, learning_rate_current)  
        all_sheet.cell(index_record, 2, batch_size)  
        all_sheet.cell(index_record, 3, epoch_id)  

        for phase in ['train', 'val', 'EVC1', 'EVC2']:
            if phase == 'train':
                model.train() 
                phase_flag = 0
            elif phase == 'val':
                model.eval()  
                phase_flag = 1
            elif phase == 'EVC1':
                model.eval()  
                phase_flag = 2
            elif phase == 'EVC2':
                model.eval()   
                phase_flag = 3

            torch.set_grad_enabled(phase == 'train')

            size = len(dataloaders[phase_flag].dataset)
            num_mini_batches = math.ceil(size / batch_size)
            print('num_mini_batches = ', num_mini_batches)

            running_loss = 0.0

            epoch_outputs_p = []
            epoch_labels_p = []

            for batch, (p_patches_batch, p_patch_LDs_batch, p_patch_SDs_batch, p_patch_RDs_batch, p_patch_adc_batch, p_LN_meta_label_batch, p_LN_meta_ratio_batch) \
                    in enumerate(dataloaders[phase_flag]):

                LN_meta_label_p_batch = torch.tensor(p_LN_meta_label_batch, dtype=torch.float)
                LN_meta_label_p_batch = LN_meta_label_p_batch.to(device)

                LN_meta_ratio_p_batch = torch.tensor(p_LN_meta_ratio_batch, dtype=torch.float)
                LN_meta_ratio_p_batch = LN_meta_ratio_p_batch.to(device)   

                img_max_outputs_p_batch = []
                img_max_outputs_p_batch = torch.tensor(img_max_outputs_p_batch)
                img_max_outputs_p_batch = img_max_outputs_p_batch.to(device)

                img_avg_outputs_p_batch = []
                img_avg_outputs_p_batch = torch.tensor(img_avg_outputs_p_batch)
                img_avg_outputs_p_batch = img_avg_outputs_p_batch.to(device)

                combined_max_outputs_p_batch = []
                combined_max_outputs_p_batch = torch.tensor(combined_max_outputs_p_batch)
                combined_max_outputs_p_batch = combined_max_outputs_p_batch.to(device)

                combined_avg_outputs_p_batch = []
                combined_avg_outputs_p_batch = torch.tensor(combined_avg_outputs_p_batch)
                combined_avg_outputs_p_batch = combined_avg_outputs_p_batch.to(device)

                img_features_max_batch = []
                img_features_max_batch = torch.tensor(img_features_max_batch)
                img_features_max_batch = img_features_max_batch.to(device)

                rest_patches_output_p_batch = []
                rest_patches_output_p_batch = torch.tensor(rest_patches_output_p_batch)
                rest_patches_output_p_batch = rest_patches_output_p_batch.to(device)

                for p_id in range(0, len(p_patches_batch)): 
                    p_patches = p_patches_batch[p_id]
                    p_patch_LDs = p_patch_LDs_batch[p_id]
                    p_patch_SDs = p_patch_SDs_batch[p_id]
                    p_patch_RDs = p_patch_RDs_batch[p_id]
                    p_patch_adcs = p_patch_adc_batch[p_id]
                    p_LN_meta_label = p_LN_meta_label_batch[p_id]
                    p_LN_meta_ratio = p_LN_meta_ratio_batch[p_id]

                    img_outputs_p_patches = []
                    img_outputs_p_patches = torch.tensor(img_outputs_p_patches)
                    img_outputs_p_patches = img_outputs_p_patches.to(device)

                    combined_outputs_p_patches = []
                    combined_outputs_p_patches = torch.tensor(combined_outputs_p_patches)
                    combined_outputs_p_patches = combined_outputs_p_patches.to(device)
                    
                    img_features = []
                    img_features = torch.tensor(img_features)
                    img_features = img_features.to(device)

                    rest_patches_output = []
                    rest_patches_output = torch.tensor(rest_patches_output)
                    rest_patches_output = rest_patches_output.to(device)

                    patch_num = len(p_patches)
                    for patch_id in range(0, patch_num):
                        current_patch, combined_patch = global_input(patch_id, p_patches)
                        current_patch = T.ToTensor()(current_patch)
                        current_patch = transform(current_patch)
                        if phase == 'train': 
                            current_patch = random_transform(current_patch)
                        current_patch = current_patch.unsqueeze(0)
                        current_patch = F.interpolate(current_patch, size=(128, 128), mode='bilinear', align_corners=False)

                        current_patch = torch.as_tensor(current_patch, dtype=torch.float)
                        current_patch = current_patch.to(device)

                        combined_patch = T.ToTensor()(combined_patch)
                        combined_patch = transform(combined_patch)
                        combined_patch = combined_patch.float()
                        combined_patch = combined_patch.unsqueeze(0)
                        combined_patch = combined_patch.to(device)

                        p_patch_LD = p_patch_LDs[patch_id]/10.0
                        p_patch_LD = torch.as_tensor(p_patch_LD, dtype=torch.float)
                        p_patch_LD = p_patch_LD.unsqueeze(0)  # 1*1
                        p_patch_LD = p_patch_LD.to(device)

                        p_patch_SD = p_patch_SDs[patch_id]/10.0  # cm
                        p_patch_SD = torch.as_tensor(p_patch_SD, dtype=torch.float)
                        p_patch_SD = p_patch_SD.unsqueeze(0)
                        p_patch_SD = p_patch_SD.to(device)

                        p_patch_RD = p_patch_RDs[patch_id]
                        p_patch_RD = torch.as_tensor(p_patch_RD, dtype=torch.float)
                        p_patch_RD = p_patch_RD.unsqueeze(0)
                        p_patch_RD = p_patch_RD.to(device)

                        p_patch_adc = p_patch_adcs[patch_id]/100.0   
                        p_patch_adc = torch.as_tensor(p_patch_adc, dtype=torch.float)
                        p_patch_adc = p_patch_adc.unsqueeze(0)
                        p_patch_adc = p_patch_adc.to(device)

                        [patch_img_pred, patch_combined_pred, logits, auxiliary_output] = model(current_patch, combined_patch , p_patch_LD, p_patch_SD, p_patch_RD, p_patch_adc)
                        combined_outputs_p_patches = torch.cat((combined_outputs_p_patches, patch_combined_pred))
                        img_features = torch.cat((img_features, logits))
                        rest_patches_output = torch.cat((rest_patches_output, auxiliary_output))

                    combined_outputs_p_patch_max, _ = torch.max(combined_outputs_p_patches, 0)
                    combined_outputs_p_patch_avg = torch.mean(combined_outputs_p_patches, 0)
                    fea_index = torch.argmax(combined_outputs_p_patches, 0)
                    img_features_max = img_features[fea_index]
                    rest_patches_output_avg = torch.mean(rest_patches_output, 0)

                    combined_max_outputs_p_batch = torch.cat(
                        (combined_max_outputs_p_batch, combined_outputs_p_patch_max))
                    combined_avg_outputs_p_batch = torch.cat(
                        (combined_avg_outputs_p_batch, combined_outputs_p_patch_avg))
                    img_features_max_batch = torch.cat(
                        (img_features_max_batch, img_features_max))
                    rest_patches_output_p_batch = torch.cat(
                        (rest_patches_output_p_batch, rest_patches_output_avg))

                loss_combined_max = criterions['loss_combined_max'](combined_max_outputs_p_batch, LN_meta_label_p_batch)
                loss_combined_avg = criterions['loss_combined_avg'](combined_avg_outputs_p_batch, LN_meta_ratio_p_batch) 
                loss_auxiliary = criterions['loss_auxiliary'](rest_patches_output_p_batch, LN_meta_ratio_p_batch)

                loss_main = loss_combined_max + 0.25*loss_combined_avg
                loss = loss_main + 0.25*loss_auxiliary

                if phase == 'train':
                    optimizer.zero_grad()
                    loss.backward()
                    optimizer.step()
                loss_item, current = loss.item(), batch * batch_size
                print(f"batch {batch+1} - loss: {loss_item:>4f}  [{current:>5d}/{size:>5d}]")

                running_loss += loss.item()
                epoch_outputs_p.extend(combined_max_outputs_p_batch.tolist())
                epoch_labels_p.extend(p_LN_meta_label_batch)

            epoch_loss = running_loss / num_mini_batches

            if phase == 'train':
                epoch_evaluations, acc_ci, auc_ci, f1_ci = evaluate(np.array(epoch_labels_p), np.array(epoch_outputs_p), digits=3)
                epoch_AUC = epoch_evaluations["AUC"]
                epoch_ACC = epoch_evaluations["ACC"]
                epoch_SEN = epoch_evaluations["SEN"]
                epoch_SPE = epoch_evaluations["SPE"]
                epoch_CUTOFF = epoch_evaluations["cutoff"]
                epoch_F1 = epoch_evaluations["F1"]
                print('{} Loss: {:.4f} {} AUC:{:.4f} {} ACC: {:.4f} {} SEN: {:.4f} {} SPE: {:.4f} {} cutoff:{:.4f} {} F1:{:.4f} {}'.format(
                    phase, epoch_loss, phase, epoch_AUC, phase, epoch_ACC, phase, epoch_SEN, phase, epoch_SPE, phase, epoch_CUTOFF, phase, epoch_F1, phase))
                all_sheet.cell(index_record, 4, epoch_loss)
                AUC_train = epoch_AUC

            if phase == 'val':
                epoch_evaluations, acc_ci, auc_ci, f1_ci = evaluate(np.array(epoch_labels_p), np.array(epoch_outputs_p), digits=3)
                epoch_AUC = epoch_evaluations["AUC"]
                epoch_ACC = epoch_evaluations["ACC"]
                epoch_SEN = epoch_evaluations["SEN"]
                epoch_SPE = epoch_evaluations["SPE"]
                epoch_CUTOFF = epoch_evaluations["cutoff"]
                epoch_F1 = epoch_evaluations["F1"]
                print('{} Loss: {:.4f} {} AUC:{:.4f} {} ACC: {:.4f} {} SEN: {:.4f} {} SPE: {:.4f} {} cutoff:{:.4f} {} F1:{:.4f}'.format(
                    phase, epoch_loss, phase, epoch_AUC, phase, epoch_ACC, phase, epoch_SEN, phase, epoch_SPE, phase, epoch_CUTOFF, phase, epoch_F1))
                all_sheet.cell(index_record, 5, epoch_AUC)
                all_sheet.cell(index_record, 8, epoch_ACC)
                all_sheet.cell(index_record, 11, epoch_F1)
                all_sheet.cell(index_record, 14, auc_ci[0])
                all_sheet.cell(index_record, 15, auc_ci[1])
                all_sheet.cell(index_record, 20, acc_ci[0])
                all_sheet.cell(index_record, 21, acc_ci[1])
                all_sheet.cell(index_record, 26, f1_ci[0])
                all_sheet.cell(index_record, 27, f1_ci[1])
                AUC_val = epoch_AUC

            if phase == 'EVC1':
                epoch_evaluations, acc_ci, auc_ci, f1_ci = evaluate(np.array(epoch_labels_p), np.array(epoch_outputs_p), digits=3)
                epoch_AUC = epoch_evaluations["AUC"]
                epoch_ACC = epoch_evaluations["ACC"]
                epoch_SEN = epoch_evaluations["SEN"]
                epoch_SPE = epoch_evaluations["SPE"]
                epoch_CUTOFF = epoch_evaluations["cutoff"]
                epoch_F1 = epoch_evaluations["F1"]
                print('{} Loss: {:.4f} {} AUC:{:.4f} {} ACC: {:.4f} {} SEN: {:.4f} {} SPE: {:.4f} {} cutoff:{:.4f} {} F1:{:.4f}'.format(
                    phase, epoch_loss, phase, epoch_AUC, phase, epoch_ACC, phase, epoch_SEN, phase, epoch_SPE, phase, epoch_CUTOFF, phase, epoch_F1))
                all_sheet.cell(index_record, 6, epoch_AUC)
                all_sheet.cell(index_record, 9, epoch_ACC)
                all_sheet.cell(index_record, 12, epoch_F1)
                all_sheet.cell(index_record, 16, auc_ci[0])
                all_sheet.cell(index_record, 17, auc_ci[1])
                all_sheet.cell(index_record, 22, acc_ci[0])
                all_sheet.cell(index_record, 23, acc_ci[1])
                all_sheet.cell(index_record, 28, f1_ci[0])
                all_sheet.cell(index_record, 29, f1_ci[1])
                AUC_EVC1 = epoch_AUC

            if phase == 'EVC2':
                epoch_evaluations, acc_ci, auc_ci, f1_ci = evaluate(np.array(epoch_labels_p), np.array(epoch_outputs_p), digits=3)
                epoch_AUC = epoch_evaluations["AUC"]
                epoch_ACC = epoch_evaluations["ACC"]
                epoch_SEN = epoch_evaluations["SEN"]
                epoch_SPE = epoch_evaluations["SPE"]
                epoch_CUTOFF = epoch_evaluations["cutoff"]
                epoch_F1 = epoch_evaluations["F1"]
                print('{} Loss: {:.4f} {} AUC:{:.4f} {} ACC: {:.4f} {} SEN: {:.4f} {} SPE: {:.4f} {} cutoff:{:.4f} {} F1:{:.4f}'.format(
                    phase, epoch_loss, phase, epoch_AUC, phase, epoch_ACC, phase, epoch_SEN, phase, epoch_SPE, phase, epoch_CUTOFF, phase, epoch_F1))
                all_sheet.cell(index_record, 7, epoch_AUC)
                all_sheet.cell(index_record, 10, epoch_ACC)
                all_sheet.cell(index_record, 13, epoch_F1)
                all_sheet.cell(index_record, 18, auc_ci[0])
                all_sheet.cell(index_record, 19, auc_ci[1])
                all_sheet.cell(index_record, 24, acc_ci[0])
                all_sheet.cell(index_record, 25, acc_ci[1])
                all_sheet.cell(index_record, 30, f1_ci[0])
                all_sheet.cell(index_record, 31, f1_ci[1])
                AUC_EVC2 = epoch_AUC

            if phase == 'val':
                scheduler.step()   

        torch.save(model.state_dict(),
                    results_save_folder +
                       'model_auc_{:.3f}_{:.3f}_{:.3f}_{:.3f}_epoch_{}.pkl'.format(AUC_train, AUC_val, AUC_EVC1, AUC_EVC2, epoch+1))
        index_record = index_record + 1

        epoch_end_time = time.time()
        print('epoch computation time: ', str(epoch_end_time-epoch_start_time))
        if epoch % 10 == 0:
            log_modelling.save(filename=log_path)
    log_modelling.save(filename=log_path)
    time_elapsed = time.time() - since
    print('Training complete in {:.0f}m {:.0f}s'.format(
        time_elapsed // 60, time_elapsed % 60))
    print('Best val AUC: {:.4f}'.format(best_val_auc))
    print()
    print()

if __name__ == '__main__':
    my_seed = 'yourseed'
    set_determinism(seed=my_seed)

    date = "date"
    data_folder_path = 'datapath'
    result_folder_path = 'datapath' 

    center1_name = ''
    center2_name = ''
    center3_name = ''

    p_patches_train, p_patch_LDs_train, p_patch_SDs_train, p_patch_RDs_train, p_patch_adc_train, p_LN_meta_label_train, p_LN_meta_ratio_train \
        = load_LN_data_by_set('', data_folder_path, 'training')
    p_patches_val, p_patch_LDs_val, p_patch_SDs_val, p_patch_RDs_val, p_patch_adc_val, p_LN_meta_label_val, p_LN_meta_ratio_val \
        = load_LN_data_by_set('', data_folder_path, 'val')
    p_patches_EVC1, p_patch_LDs_EVC1, p_patch_SDs_EVC1, p_patch_RDs_EVC1, p_patch_adc_EVC1, p_LN_meta_label_EVC1, p_LN_meta_ratio_EVC1 \
        = load_LN_data('', data_folder_path)
    p_patches_EVC2, p_patch_LDs_EVC2, p_patch_SDs_EVC2, p_patch_RDs_EVC2, p_patch_adc_EVC2, p_LN_meta_label_EVC2, p_LN_meta_ratio_EVC2 \
        = load_LN_data('', data_folder_path)

    device = torch.device("cuda:")
    batch_size_list = [8] 
    learning_rate_list = [4e-4] 
    epoch = 200
    ra_loss_weight = 0.5
    ram_start = 0

    loss_fuc = 'BCELoss'
    criterions = OrderedDict()
    criterions['loss_img_max'] = nn.BCELoss() 
    criterions['loss_img_avg'] = nn.MSELoss() 
    criterions['loss_combined_max'] = nn.BCELoss()  
    criterions['loss_combined_avg'] = nn.MSELoss() 
    criterions['loss_auxiliary'] = nn.MSELoss()   

    model_name = 'WeGA'

    work_dir = result_folder_path + model_name + '_' + loss_fuc + '_externalTest' + '_date_' + date + '/'
    if not os.path.exists(work_dir):
        os.makedirs(work_dir)
    log_path = work_dir + '/log.xlsx'
    log_modelling = log_create()
    log_modelling.save(filename=log_path)
    index_record = 2

    for batch_size in batch_size_list:
        for learning_rate in learning_rate_list:
            print(
                "Number of samples in train, validation, EVC1, and EVC2 are %d, %d, %d and %d."
                % (len(p_patches_train), len(p_patches_val), len(p_patches_EVC1), len(p_patches_EVC2)))

            train_dataset = LN_Dataset(p_T2_patch_list=p_patches_train,
                                       p_patch_LD_list=p_patch_LDs_train,
                                       p_patch_SD_list=p_patch_SDs_train,
                                       p_patch_RD_list=p_patch_RDs_train,
                                       p_patch_adc_list=p_patch_adc_train,
                                       p_LN_meta_label_list=p_LN_meta_label_train,
                                       p_LN_meta_ratio=p_LN_meta_ratio_train)
            train_dataloader = torch.utils.data.DataLoader(train_dataset, batch_size=batch_size, shuffle=True,
                                                           num_workers=4, pin_memory=True, collate_fn=my_collate)

            val_dataset = LN_Dataset(p_T2_patch_list=p_patches_val,
                                     p_patch_LD_list=p_patch_LDs_val,
                                     p_patch_SD_list=p_patch_SDs_val,
                                     p_patch_RD_list=p_patch_RDs_val,
                                     p_patch_adc_list=p_patch_adc_val,
                                     p_LN_meta_label_list=p_LN_meta_label_val,
                                     p_LN_meta_ratio=p_LN_meta_ratio_val)
            val_dataloader = torch.utils.data.DataLoader(val_dataset, batch_size=batch_size,
                                                           num_workers=4, pin_memory=True, collate_fn=my_collate)

            EVC1_dataset = LN_Dataset(p_T2_patch_list=p_patches_EVC1,
                                     p_patch_LD_list=p_patch_LDs_EVC1,
                                     p_patch_SD_list=p_patch_SDs_EVC1,
                                     p_patch_RD_list=p_patch_RDs_EVC1,
                                     p_patch_adc_list=p_patch_adc_EVC1,
                                     p_LN_meta_label_list=p_LN_meta_label_EVC1,
                                     p_LN_meta_ratio=p_LN_meta_ratio_EVC1)
            EVC1_dataloader = torch.utils.data.DataLoader(EVC1_dataset, batch_size=batch_size,
                                                           num_workers=4, pin_memory=True, collate_fn=my_collate)


            EVC2_dataset = LN_Dataset(p_T2_patch_list=p_patches_EVC2,
                                     p_patch_LD_list=p_patch_LDs_EVC2,
                                     p_patch_SD_list=p_patch_SDs_EVC2,
                                     p_patch_RD_list=p_patch_RDs_EVC2,
                                     p_patch_adc_list=p_patch_adc_EVC2,
                                     p_LN_meta_label_list=p_LN_meta_label_EVC2,
                                     p_LN_meta_ratio=p_LN_meta_ratio_EVC2)
            EVC2_dataloader = torch.utils.data.DataLoader(EVC2_dataset, batch_size=batch_size,
                                                           num_workers=4, pin_memory=True, collate_fn=my_collate)

            dataloaders = [train_dataloader, val_dataloader, EVC1_dataloader, EVC2_dataloader]
            model_ft = Combined_network(1)
            model_ft = model_ft.to(device)

            optimizer_ft = optim.AdamW(model_ft.parameters(), lr=learning_rate)
            exp_lr_scheduler = lr_scheduler.ReduceLROnPlateau(optimizer_ft, mode='min', factor=0.5, patience=10, threshold=0.001,
                                                       threshold_mode='abs', cooldown=5, min_lr=1e-7, eps=1e-08, verbose=False)

            CosineAnnealingLR_scheduler = lr_scheduler.CosineAnnealingLR(optimizer_ft, T_max=(epoch * len(
                p_LN_meta_label_train)) // batch_size / 10, eta_min=1e-8, verbose=True)


            torch.backends.cudnn.benchmark = True
            train_model(device, dataloaders, model_ft, criterions,
                        optimizer_ft, CosineAnnealingLR_scheduler, batch_size,
                        learning_rate, epoch, work_dir, log_path, log_modelling, index_record, ra_loss_weight, ram_start)
            index_record = index_record + epoch

